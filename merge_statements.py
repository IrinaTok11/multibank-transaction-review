"""
Merge bank statements from different bank exports into one Excel table.

The script detects the header row, maps columns, normalizes dates and amounts,
and writes the merged result to merged_statements.xlsx.
"""
import argparse
import hashlib
import re
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

# =============================================================================
# COLUMN MAPPING CONFIG
# =============================================================================
BANK_COLUMN_MAPS = {

    # Two separate Debit / Credit columns
    "DebitCredit": {
        "Date":                       "date",
        "Description of transaction": "counterparty",
        "Debit":                      "debit",
        "Credit":                     "credit",
    },

    # Single signed Amount column — split automatically
    "SignedAmount": {
        "Date":        "date",
        "Description": "counterparty",
        "Amount":      "amount",   # positive -> credit, negative -> debit
    },

    # Paid in / Paid out columns
    "PaidInOut": {
        "Date":           "date",
        "Transaction":    "counterparty",
        "Paid in (USD)":  "credit",
        "Paid out (USD)": "debit",
    },

    # Debit / Credit columns with transaction date header
    "DebitCredit_Extended": {
        "Transaction Date": "date",
        "Description":      "counterparty",
        "Debit":            "debit",
        "Credit":           "credit",
    },

    # Withdrawals / Deposits columns
    "WithdrawalDeposit": {
        "Date":        "date",
        "Description": "counterparty",
        "Withdrawals": "debit",
        "Deposits":    "credit",
    },

    # Split counterparty: separate Payee (debit) and Payer (credit)
    "PayeePayerSplit": {
        "Date":              "date",
        "Payee":             "counterparty_debit",
        "Payer":             "counterparty_credit",
        "Withdrawal Amount": "debit",
        "Deposit Amount":    "credit",
    },
}

# =============================================================================
# UTILITY
# =============================================================================

def _normalize_col(s: str) -> str:
    """Normalize a column name for tolerant matching across bank exports."""
    return re.sub(r"\s+", " ", str(s).strip().lower().replace("_", " "))

_MONTH_MAP = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}

def normalize_date_column(series: pd.Series) -> pd.Series:
    """Convert dates from typical bank exports to pandas datetime.
    
    Handles common US/EU formats and normalizes month abbreviations.
    Ambiguous generic dates raise ValueError.
    """
    s = series.astype("string").str.strip()

    result = pd.to_datetime(s, format="%m/%d/%Y", errors="coerce")
    result = result.combine_first(pd.to_datetime(s, format="%d.%m.%Y", errors="coerce"))

    # Convert month abbreviations to numbers so date parsing does not depend on system locale.
    s_month_norm = s.str.replace(
        r"(?i)\b([A-Za-z]{3})\b",
        lambda m: _MONTH_MAP.get(m.group(1).lower(), m.group(1)),
        regex=True,
    )
    result = result.combine_first(
        pd.to_datetime(s_month_norm, format="%d-%m-%Y", errors="coerce")
    )

    # Re-check only the values that still were not parsed as dates.
    remaining = s[result.isna()]
    if not remaining.empty:
        fallback_dayfirst = pd.to_datetime(remaining, dayfirst=True, errors="coerce")
        fallback_monthfirst = pd.to_datetime(remaining, dayfirst=False, errors="coerce")

        ambiguous_mask = (
            fallback_dayfirst.notna()
            & fallback_monthfirst.notna()
            & (fallback_dayfirst != fallback_monthfirst)
        )
        if ambiguous_mask.any():
            examples = remaining[ambiguous_mask].head(3).tolist()
            raise ValueError(
                f"Ambiguous date format detected — could be day-first or month-first, "
                f"e.g. {examples}. Review these values and use one explicit date format."
            )

        result = result.combine_first(fallback_dayfirst)
        result = result.combine_first(fallback_monthfirst)

    return result

def normalize_amount_column(series: pd.Series) -> pd.Series:
    """Convert amount strings from common bank export formats to floats.

    Supports US/EU separators, negative values in parentheses, trailing minus,
    unicode minus, and dash decimals such as 1000-00.
    """
    s = series.astype("string").str.strip()

    # Replace the unicode minus with the standard minus sign.
    s = s.str.replace("\u2212", "-", regex=False)

    # Convert a trailing minus into a regular negative amount: "123.45-" -> "-123.45".
    trailing_neg = s.str.endswith("-") & ~s.str.startswith("-")
    s = s.where(~trailing_neg, "-" + s.str.slice(0, -1))

    # Treat amounts in parentheses as negative: "(1,234.56)" -> "-1,234.56".
    is_parens = s.str.startswith("(") & s.str.endswith(")")
    s = s.where(~is_parens, "-" + s.str.slice(1, -1))

    # Leave only digits, separators, and the minus sign; currency labels such as "USD" are removed.
    s = s.str.replace(r"[^0-9,\.\-]", "", regex=True)

    is_neg = s.str.startswith("-")
    s = s.str.replace(r"^-", "", regex=True)

    # Remove dot thousand separators in values like "1.234" or "1.234.567".
    mask_thousand_dot = s.str.match(r"^\d{1,3}(\.\d{3})+$")
    s = s.where(~mask_thousand_dot, s.str.replace(".", "", regex=False))

    # For European amounts, remove thousand dots before converting the decimal comma.
    s = s.str.replace(r"\.(\d{3})(?=[,\d])", r"\1", regex=True)

    # Treat a comma as the decimal separator only when two decimal digits follow, as in bank amounts.
    s = s.str.replace(r",(\d{2})$", r".\1", regex=True)

    # Any commas left at this stage are treated as thousand separators.
    s = s.str.replace(",", "", regex=False)

    # Convert forms like "1000-00" into a normal decimal amount.
    s = s.str.replace(r"-(\d{2})$", r".\1", regex=True)

    result = pd.to_numeric(s, errors="coerce")
    result = result.where(~is_neg, -result)

    return result

def split_signed_amount(df: pd.DataFrame) -> pd.DataFrame:
    """SignedAmount format: split a single signed amount column into debit/credit.

    NaN amounts are preserved as NaN (not forced to 0) to avoid masking parse errors.
    Positive amounts -> credit, negative amounts -> debit (as positive).
    """
    if "amount" not in df.columns:
        return df

    amt = pd.to_numeric(df["amount"], errors="coerce")
    df["credit"] = amt.where(amt > 0)
    df["debit"] = (-amt).where(amt < 0)

    df.drop(columns="amount", inplace=True)
    return df

def make_txn_id(row: pd.Series, index: int) -> str:
    """Create a stable transaction id from the key transaction fields.

    Uses date, debit, credit, counterparty, source file, and row index.
    Empty counterparty values are treated as an empty string.
    """
    date_val = row.get("date", "")
    if pd.notna(date_val) and hasattr(date_val, "date"):
        date_str = date_val.date().isoformat()
    else:
        date_str = str(date_val)

    debit_val = pd.to_numeric(row.get("debit", 0), errors="coerce")
    credit_val = pd.to_numeric(row.get("credit", 0), errors="coerce")
    debit = 0.0 if pd.isna(debit_val) else float(debit_val)
    credit = 0.0 if pd.isna(credit_val) else float(credit_val)

    debit_str = f"{debit:.4f}"
    credit_str = f"{credit:.4f}"

    cp = row.get("counterparty", "")
    cp_str = "" if pd.isna(cp) else str(cp)

    source = row.get("source_file", "")
    key = f"{date_str}{debit_str}{credit_str}{cp_str}{source}{index}"

    return hashlib.md5(key.encode("utf-8")).hexdigest()[:12].upper()

def find_header_row(path: Path, max_scan: int = 30) -> int:
    """Return the 0-based header row index with the best column-name match score."""
    raw = load_file(path, header_row=None).head(max_scan)

    all_expected = {
        _normalize_col(c)
        for mapping in BANK_COLUMN_MAPS.values()
        for c in mapping.keys()
    }

    best_row, best_score = 0, 0
    for i, row in raw.iterrows():
        cells = {
            _normalize_col(str(v))
            for v in row.dropna().values
            if str(v).strip()
        }
        score = len(cells & all_expected)
        if score > best_score:
            best_score, best_row = score, i

    if best_score == 0:
        raise ValueError(
            f"{path.name}: could not detect a header row in the first {max_scan} rows"
        )

    return int(best_row)

# =============================================================================
# COLUMN DETECTION AND NORMALIZATION
# =============================================================================

def detect_format_and_map(df: pd.DataFrame) -> tuple[str, dict]:
    """Detect the statement format by the file columns.

    Requires date, description, and amount information. If more than one format
    fits, the best match is used.
    """
    raw_cols_norm = {_normalize_col(c): c for c in df.columns}
    best_fmt, best_score = None, 0.0

    for fmt, mapping in BANK_COLUMN_MAPS.items():
        mapped_norm = {_normalize_col(k): v for k, v in mapping.items()}

        date_keys = {k for k, v in mapped_norm.items() if v == "date"}
        amount_keys = {k for k, v in mapped_norm.items() if v in {"debit", "credit", "amount"}}
        cp_keys = {k for k, v in mapped_norm.items() if v in {"counterparty", "counterparty_debit", "counterparty_credit"}}

        if not (date_keys & raw_cols_norm.keys()):
            continue
        if not (amount_keys & raw_cols_norm.keys()):
            continue
        if not (cp_keys & raw_cols_norm.keys()):
            continue

        score = sum(1 for k in mapped_norm if k in raw_cols_norm) / len(mapped_norm)
        if score < 0.4:
            continue
        if score > best_score:
            best_score, best_fmt = score, fmt

    if best_fmt is None:
        raise ValueError(
            "Could not identify format. Check header row detection and BANK_COLUMN_MAPS."
        )

    return best_fmt, BANK_COLUMN_MAPS[best_fmt]

def apply_column_map(df: pd.DataFrame, column_map: dict) -> pd.DataFrame:
    raw_norm = {_normalize_col(c): c for c in df.columns}
    rename = {
        raw_norm[_normalize_col(raw)]: canonical
        for raw, canonical in column_map.items()
        if _normalize_col(raw) in raw_norm
    }
    df = df.rename(columns=rename)

    ordered: list[str] = []
    for c in rename.values():
        if c in df.columns and c not in ordered:
            ordered.append(c)

    return df[ordered].copy()

def choose_counterparty(
    payee: pd.Series,
    payer: pd.Series,
    is_expense: pd.Series,
) -> pd.Series:
    """Use payee for debit rows and payer for credit rows, with blank fallback."""
    result = payer.copy()
    result[is_expense] = payee[is_expense].where(payee[is_expense].ne(""), payer[is_expense])
    result[~is_expense] = payer[~is_expense].where(payer[~is_expense].ne(""), payee[~is_expense])
    return result

def resolve_counterparty(df: pd.DataFrame) -> pd.DataFrame:
    """PayeePayerSplit format: merge counterparty_debit / counterparty_credit into one column."""
    has_split = "counterparty_debit" in df.columns or "counterparty_credit" in df.columns
    if not has_split:
        if "counterparty" not in df.columns:
            df["counterparty"] = ""
        return df

    debit = pd.to_numeric(
        df.get("debit", pd.Series(0, index=df.index)),
        errors="coerce",
    ).fillna(0)
    credit = pd.to_numeric(
        df.get("credit", pd.Series(0, index=df.index)),
        errors="coerce",
    ).fillna(0)

    both_nonzero = (debit > 0) & (credit > 0)
    if both_nonzero.any():
        raise ValueError("Rows with both debit and credit > 0 detected in PayeePayerSplit format")

    payee = df.get("counterparty_debit", pd.Series("", index=df.index)).fillna("").astype("string").str.strip()
    payer = df.get("counterparty_credit", pd.Series("", index=df.index)).fillna("").astype("string").str.strip()
    is_expense = debit > 0

    df["counterparty"] = choose_counterparty(payee, payer, is_expense)
    for col in ("counterparty_debit", "counterparty_credit"):
        df.drop(columns=col, errors="ignore", inplace=True)
    return df

# =============================================================================
# FILE LOADING
# =============================================================================

def load_file(path: Path, header_row: int = 0) -> pd.DataFrame:
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return pd.read_excel(path, header=header_row, dtype=str, engine="openpyxl")

    if suffix == ".xls":
        try:
            return pd.read_excel(path, header=header_row, dtype=str, engine="xlrd")
        except ImportError as exc:
            raise RuntimeError(
                f"Reading '{path.name}' requires the 'xlrd' package. "
                "Install it with: pip install xlrd   or convert the file to .xlsx."
            ) from exc

    raise ValueError(f"Unsupported file type: {path.suffix}")

def load_and_normalize(path: Path) -> pd.DataFrame:
    header_row = find_header_row(path, max_scan=30)
    print(f"header row: {header_row + 1} (1-based)")

    raw = load_file(path, header_row=header_row)
    raw.dropna(how="all", inplace=True)

    fmt_name, col_map = detect_format_and_map(raw)
    df = apply_column_map(raw, col_map)

    if "date" in df.columns:
        raw_date = df["date"].astype("string")
        df["date"] = normalize_date_column(df["date"])
        bad_date = df["date"].isna() & raw_date.str.strip().fillna("").ne("")
        if bad_date.any():
            examples = raw_date[bad_date].head(3).tolist()
            raise ValueError(f"{path.name}: {bad_date.sum()} unparsed dates, e.g. {examples}")

    for col in ("debit", "credit", "amount"):
        if col in df.columns:
            raw_amt = df[col].astype("string")
            df[col] = normalize_amount_column(df[col])
            bad_amt = df[col].isna() & raw_amt.str.strip().fillna("").ne("")
            if bad_amt.any():
                examples = raw_amt[bad_amt].head(3).tolist()
                raise ValueError(
                    f"{path.name}: {bad_amt.sum()} unparsed amounts in '{col}', e.g. {examples}"
                )

    if fmt_name == "SignedAmount":
        df = split_signed_amount(df)

    if "debit" in df.columns and "credit" in df.columns:
        df = df[(df["debit"].fillna(0).abs() + df["credit"].fillna(0).abs()) > 0].copy()

    df = resolve_counterparty(df)

    if df.empty:
        raise ValueError(f"{path.name}: no transactions after cleaning (header detection or parsing issue)")

    if "date" in df.columns and df["date"].isna().all():
        raise ValueError(f"{path.name}: all dates are NaT after parsing")

    if "counterparty" in df.columns:
        if df["counterparty"].astype("string").str.strip().fillna("").eq("").all():
            raise ValueError(f"{path.name}: all descriptions are empty after mapping")

    df["fmt"] = fmt_name
    df["source_file"] = path.name

    return df

def format_transactions_sheet(ws, df: pd.DataFrame) -> None:
    col_widths = {
        "txn_id": 14, "Date": 14, "Description": 50,
        "Debit": 15, "Credit": 15, "Source_File": 36,
    }
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "A2"
    for idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col_widths.get(col_name, 18)
    if "Date" in df.columns:
        date_letter = get_column_letter(list(df.columns).index("Date") + 1)
        # Set the Excel date format explicitly so the displayed date stays the same on different systems.
        for row in range(2, len(df) + 2):
            ws[f"{date_letter}{row}"].number_format = "DD/MM/YYYY"
    for col_name in ("Debit", "Credit"):
        if col_name in df.columns:
            col_letter = get_column_letter(list(df.columns).index(col_name) + 1)
            for row in range(2, len(df) + 2):
                ws[f"{col_letter}{row}"].number_format = "#,##0.00"

# =============================================================================
# MAIN
# =============================================================================

def merge_statements(input_dir: Path, strict: bool = False) -> tuple[pd.DataFrame, list[tuple[str, str]]]:
    """Load and normalize all statement files from a directory.

    Returns the merged dataframe and a list of files that failed with an error.
    """
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")

    statement_files = sorted(
        f
        for f in input_dir.iterdir()
        if f.suffix.lower() in {".xlsx", ".xls"}
        and not f.name.startswith("~")
        and not f.name.startswith(".~")
    )
    if not statement_files:
        raise ValueError(f"No statement files found in: {input_dir}")

    all_frames: list[pd.DataFrame] = []
    failed: list[tuple[str, str]] = []

    for path in statement_files:
        try:
            print(f"  Loading {path.name} ...")
            df = load_and_normalize(path)

            # Keep the row number from the original file so txn_id stays stable across runs.
            df["_row_in_file"] = range(len(df))

            all_frames.append(df)
            fmt = df["fmt"].iat[0]
            print(f"  OK  {path.name}  ({len(df):,} rows, format={fmt})")
        except (ValueError, RuntimeError, OSError) as exc:
            msg = str(exc)
            print(f"  ERR {path.name}: {msg}")
            failed.append((path.name, msg))
            if strict:
                raise

    if not all_frames:
        raise ValueError("No files loaded.")

    result = pd.concat(all_frames, ignore_index=True, sort=False)
    result.attrs["files_total"] = len(statement_files)
    result.attrs["files_processed"] = len(all_frames)

    required = {"date", "counterparty", "debit", "credit", "fmt", "source_file"}
    missing = required - set(result.columns)
    if missing:
        raise RuntimeError(f"Internal error: missing columns after concat: {sorted(missing)}")

    return result, failed

def generate_txn_ids(result: pd.DataFrame) -> pd.Series:
    """Generate deterministic transaction ids, with a row-wise fallback for unexpected dtypes."""
    try:
        idx = result["_row_in_file"].fillna(result.index).astype(int).astype(str)

        date_part = pd.to_datetime(result["date"], errors="coerce").dt.date.astype("string").fillna("")
        debit_part = pd.to_numeric(result["debit"], errors="coerce").fillna(0).map(lambda x: f"{x:.4f}")
        credit_part = pd.to_numeric(result["credit"], errors="coerce").fillna(0).map(lambda x: f"{x:.4f}")
        cp_part = result["counterparty"].astype("string").fillna("").str.strip()
        src_part = result["source_file"].astype("string").fillna("").str.strip()

        keys = date_part + debit_part + credit_part + cp_part + src_part + idx
        # MD5 is used here only to build a short stable row identifier; this is not a security use case.
        return keys.map(lambda s: hashlib.md5(s.encode("utf-8")).hexdigest()[:12].upper())
    except (TypeError, ValueError, AttributeError):
        return result.apply(
            lambda row: make_txn_id(row, int(row.get("_row_in_file", row.name))),
            axis=1,
        )

def build_output_dataframe(result: pd.DataFrame) -> pd.DataFrame:
    """Build the final user-facing output dataframe with friendly column names."""
    out = result[["txn_id", "date", "counterparty", "debit", "credit", "source_file"]].rename(
        columns={
            "date": "Date",
            "counterparty": "Description",
            "debit": "Debit",
            "credit": "Credit",
            "source_file": "Source_File",
        }
    )

    out["Debit"] = pd.to_numeric(out["Debit"], errors="coerce").fillna(0)
    out["Credit"] = pd.to_numeric(out["Credit"], errors="coerce").fillna(0)
    return out

def write_output_excel(out: pd.DataFrame, output_path: Path) -> None:
    """Write the consolidated output workbook to disk."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        output_path.unlink()
    except FileNotFoundError:
        pass
    except PermissionError as exc:
        print(f"Cannot overwrite '{output_path}': close the file in Excel and retry.")
        raise SystemExit(1) from exc

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl", date_format="DD/MM/YYYY") as writer:
            out.to_excel(writer, sheet_name="Transactions", index=False)
            format_transactions_sheet(writer.sheets["Transactions"], out)
    except PermissionError as exc:
        print(f"Cannot write '{output_path}': close the file in Excel and retry.")
        raise SystemExit(1) from exc

def print_run_summary(
    result: pd.DataFrame,
    out: pd.DataFrame,
    failed: list[tuple[str, str]],
    output_path: Path,
) -> None:
    """Print a compact run summary for the console."""
    print(f"\n{'=' * 60}")
    files_total = int(getattr(result, "attrs", {}).get("files_total", 0) or 0)
    files_processed = int(getattr(result, "attrs", {}).get("files_processed", 0) or 0)
    print(f"Files processed : {files_processed}/{files_total}")
    print(f"Total rows      : {len(out):,}")
    print(f"Output saved to : {output_path.resolve()}")
    if failed:
        names = ", ".join(name for name, _ in failed)
        print(f"Files skipped   : {len(failed)} ({names})")
    print(f"{'=' * 60}\n")

def run(input_dir: Path, output_path: Path, strict: bool) -> None:
    """Run the full merge pipeline from input folder to output workbook."""
    result, failed = merge_statements(input_dir=input_dir, strict=strict)
    result = result.sort_values(
        ["date", "source_file", "_row_in_file"],
        na_position="last",
    ).reset_index(drop=True)

    result.insert(0, "txn_id", generate_txn_ids(result))
    out = build_output_dataframe(result)
    write_output_excel(out, output_path)
    print_run_summary(result, out, failed, output_path)

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Merge bank statement files into a consolidated Excel workbook."
    )

    script_dir = Path(__file__).parent
    default_input = script_dir / "input"
    default_output = script_dir / "merged_statements.xlsx"

    parser.add_argument(
        "--input", "-i", default=str(default_input),
        help=f"Folder with statement files (default: {default_input})",
    )
    parser.add_argument(
        "--output", "-o", default=str(default_output),
        help=f"Output file (default: {default_output})",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Abort on the first file-level error instead of skipping failed files.",
    )
    args = parser.parse_args()

    input_dir = Path(args.input)
    output_path = Path(args.output)

    try:
        run(input_dir=input_dir, output_path=output_path, strict=args.strict)
    except Exception as exc:
        print(str(exc))
        raise SystemExit(1) from exc

if __name__ == "__main__":
    main()
